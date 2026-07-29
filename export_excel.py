"""Exportacion a `.xlsx` de existencias e historial de ventas, a demanda (CLI-06).

Esto **no** es volver a Excel. El libro que produce `exportar_a_excel` es un
reporte de salida puntual: no forma parte del guardado diario, no se lee de
vuelta y no toca el `.db`. La fuente de verdad sigue siendo SQLite (ADR-4).

El modulo **no ejecuta SQL**: lee exclusivamente a traves de la capa de dominio
(`core_existencias.obtener_existencias` y `core_historial.obtener_ventas_historial`).
Se consumen los SUBMODULOS y no la fachada `core`, porque `core` re-exporta este
modulo: importarla desde aqui cerraria un ciclo que hoy solo no revienta por el
orden alfabetico de los imports en los consumidores.
Su unica escritura es el archivo destino.

**Contrato de columnas (desviacion D10).** La hoja "Ventas" se declara como pares
`(titulo del .xlsx, clave del dict)` en `COLUMNAS_VENTAS` porque el historial de
CLI-05 entrega claves en minusculas (`precio_costo`, `total_pagado`, ...) que no
coinciden con los titulos legibles. Leer por titulo produciria una hoja con
encabezados correctos y **todas las celdas vacias** -- un fallo silencioso. La
suite compara las claves de `COLUMNAS_VENTAS` contra `core_historial.CAMPOS_HISTORIAL`,
de modo que un cambio del contrato rompe la exportacion en vez de vaciarla.

Dos ausencias deliberadas respecto del plan original:

* `"Forma de pago"` -- desde CLI-03 los pagos son N por venta, asi que una sola
  columna ya no modela la realidad. En su lugar van `Pagado` y `Saldo`, que es
  lo que el historial si expone (`total_pagado` / `saldo_pendiente`).
* `"Observaciones"` -- no esta en `CAMPOS_HISTORIAL`; traerla exigiria otra
  consulta, y este modulo no consulta la base.
"""

from __future__ import annotations

import logging
import sqlite3
from typing import Any, Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import core_existencias
import core_historial
from core_comun import CoreError

logger: Final[logging.Logger] = logging.getLogger(__name__)

#: Nombres de las dos hojas del libro, en el orden en que se crean.
HOJA_EXISTENCIAS: Final[str] = "Existencias"
HOJA_VENTAS: Final[str] = "Ventas"

#: Pares `(titulo de la columna, clave del dict de dominio)`.
COLUMNAS_EXISTENCIAS: Final[tuple[tuple[str, str], ...]] = (
    ("Codigo articulo", "Codigo articulo"),
    ("Descripcion", "Descripcion"),
    ("Piezas recibidas", "Piezas recibidas"),
    ("Piezas vendidas", "Piezas vendidas"),
    ("Piezas disponibles", "Piezas disponibles"),
    ("Precio unitario costo", "Precio unitario costo"),
    ("Total pagado real", "Total pagado real"),
    ("Valor catalogo total", "Valor catalogo total"),
)

#: Encabezado corregido de la hoja "Ventas" (D10): cada clave existe de verdad
#: en `core_historial.CAMPOS_HISTORIAL`.
COLUMNAS_VENTAS: Final[tuple[tuple[str, str], ...]] = (
    ("Fecha", "fecha"),
    ("Cliente", "cliente"),
    ("Codigo", "codigo"),
    ("Descripcion", "descripcion"),
    ("Cantidad", "cantidad"),
    ("Precio costo", "precio_costo"),
    ("Precio publico", "precio_publico"),
    ("Total", "total"),
    ("Ganancia", "ganancia"),
    ("Pagado", "total_pagado"),
    ("Saldo", "saldo_pendiente"),
)

#: Titulos cuyas celdas de datos llevan formato de moneda.
MONEDA_EXISTENCIAS: Final[frozenset[str]] = frozenset(
    {"Precio unitario costo", "Total pagado real", "Valor catalogo total"}
)
MONEDA_VENTAS: Final[frozenset[str]] = frozenset(
    {"Precio costo", "Precio publico", "Total", "Ganancia", "Pagado", "Saldo"}
)

#: Estilo de referencia del libro maestro de la epoca Excel (`formatear_hoja`).
COLOR_ENCABEZADO: Final[str] = "12C1B4"
COLOR_FUENTE_ENCABEZADO: Final[str] = "FFFFFF"
FORMATO_MONEDA: Final[str] = '"$"#,##0.00'
ANCHO_MINIMO: Final[int] = 12
ANCHO_MAXIMO: Final[int] = 45
ANCHO_HOLGURA: Final[int] = 2

_MSG_ESCRITURA: Final[str] = "No se pudo escribir el archivo de exportacion {ruta}: {detalle}"


class ExportError(CoreError):
    """Error de dominio de la exportacion a Excel."""


def _titulos(columnas: tuple[tuple[str, str], ...]) -> list[str]:
    """Encabezado de una hoja a partir de su tabla de columnas.

    Time: O(c) sobre el numero de columnas | Space: O(c)
    """
    return [titulo for titulo, _ in columnas]


def _fila(registro: dict, columnas: tuple[tuple[str, str], ...]) -> list[Any]:
    """Proyecta un dict de dominio sobre las columnas, leyendo por clave.

    Se usa `.get` a proposito: una clave ausente deja la celda vacia en vez de
    abortar la exportacion completa.

    Time: O(c) sobre el numero de columnas | Space: O(c)
    """
    return [registro.get(clave) for _, clave in columnas]


def _escribir_hoja(
    hoja: Worksheet,
    columnas: tuple[tuple[str, str], ...],
    registros: list[dict],
) -> None:
    """Vuelca el encabezado y una fila por registro, en el orden recibido.

    Time: O(f * c) sobre filas y columnas | Space: O(c)
    """
    hoja.append(_titulos(columnas))
    for registro in registros:
        hoja.append(_fila(registro, columnas))


def _estilizar_encabezado(hoja: Worksheet) -> None:
    """Aplica relleno solido, fuente blanca en negrita y centrado a la fila 1.

    Time: O(c) sobre el numero de columnas | Space: O(1)
    """
    relleno = PatternFill(
        start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid"
    )
    fuente = Font(bold=True, color=COLOR_FUENTE_ENCABEZADO)
    alineacion = Alignment(horizontal="center", vertical="center")
    for celda in hoja[1]:
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = alineacion


def _ancho_columna(celdas: tuple[Any, ...]) -> int:
    """Ancho de una columna: el texto mas largo mas holgura, acotado a [12, 45].

    Time: O(f) sobre el numero de filas | Space: O(1)
    """
    largo = max((len(str(celda.value)) if celda.value is not None else 0) for celda in celdas)
    return min(max(largo + ANCHO_HOLGURA, ANCHO_MINIMO), ANCHO_MAXIMO)


def _aplicar_moneda(celdas: tuple[Any, ...]) -> None:
    """Marca como moneda las celdas de datos de una columna (omite el titulo).

    Time: O(f) sobre el numero de filas | Space: O(1)
    """
    for celda in celdas[1:]:
        celda.number_format = FORMATO_MONEDA


def _formatear_hoja(hoja: Worksheet, columnas_moneda: frozenset[str]) -> None:
    """Estiliza el encabezado, ajusta anchos y aplica el formato de moneda.

    Time: O(f * c) sobre filas y columnas | Space: O(1)
    """
    _estilizar_encabezado(hoja)
    for celdas in hoja.columns:
        hoja.column_dimensions[get_column_letter(celdas[0].column)].width = _ancho_columna(celdas)
        if celdas[0].value in columnas_moneda:
            _aplicar_moneda(celdas)


def _construir_libro(existencias: list[dict], ventas: list[dict]) -> Workbook:
    """Arma el libro de dos hojas ya escritas y formateadas.

    Time: O(f * c) sobre filas y columnas | Space: O(f * c)
    """
    libro = Workbook()
    hoja_existencias = libro.active
    hoja_existencias.title = HOJA_EXISTENCIAS
    _escribir_hoja(hoja_existencias, COLUMNAS_EXISTENCIAS, existencias)
    _formatear_hoja(hoja_existencias, MONEDA_EXISTENCIAS)

    hoja_ventas = libro.create_sheet(HOJA_VENTAS)
    _escribir_hoja(hoja_ventas, COLUMNAS_VENTAS, ventas)
    _formatear_hoja(hoja_ventas, MONEDA_VENTAS)
    return libro


def exportar_a_excel(conn: sqlite3.Connection, ruta_destino: str) -> str:
    """Exporta existencias e historial de ventas a un `.xlsx` (R1-R8).

    El libro lleva exactamente dos hojas, `"Existencias"` y `"Ventas"`, con el
    encabezado estilizado, formato de moneda en las columnas de dinero y ancho
    automatico acotado a [12, 45]. Con la base vacia se genera igualmente el
    libro, con las dos hojas y solo sus encabezados. La base no se modifica: se
    lee via `core`, nunca con SQL propio.

    Args:
        conn: conexion inyectada desde el call-site (`db.get_conn`).
        ruta_destino: ruta del `.xlsx` a escribir.

    Returns:
        `ruta_destino`, para encadenar con el aviso de la GUI.
    Raises:
        ExportError: si el archivo destino no puede escribirse.
        CoreError: si falla la lectura de dominio (`VentaError` incluida).

    Time: O(f * c) sobre filas y columnas | Space: O(f * c)
    """
    existencias = core_existencias.obtener_existencias(conn)
    ventas = core_historial.obtener_ventas_historial(conn)
    libro = _construir_libro(existencias, ventas)

    try:
        libro.save(ruta_destino)
    except OSError as exc:
        raise ExportError(_MSG_ESCRITURA.format(ruta=ruta_destino, detalle=exc)) from exc

    logger.info(
        "Exportacion generada en %s (%d existencias, %d lineas de venta)",
        ruta_destino,
        len(existencias),
        len(ventas),
    )
    return ruta_destino
