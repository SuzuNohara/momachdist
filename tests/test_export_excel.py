"""Suite de la exportacion a `.xlsx` (CLI-06, R1-R8).

La fixture levanta el esquema real con `db.init_db(":memory:")` y siembra las
ventas con `core.registrar_venta` / `core.agregar_pago`, de modo que el
historial que se exporta es el que produce la capa de dominio de verdad -- no
un doble. El libro se escribe siempre en `tmp_path` y se vuelve a leer con
`load_workbook`, asi que las aserciones miran el archivo, no el objeto en
memoria.

**Guard central (desviacion D10).** Un test que solo compare el encabezado de la
hoja "Ventas" pasaria aunque cada celda saliera vacia, que es exactamente lo que
ocurriria leyendo el historial por titulo legible en vez de por la clave real de
`core.CAMPOS_HISTORIAL`. Por eso hay dos aserciones adicionales:
`test_columnas_ventas_*_contrato_historial` (las claves declaradas existen en el
contrato) y `test_hoja_ventas_puebla_todas_las_celdas_*` (ninguna celda de datos
queda vacia con una venta sembrada).
"""

from __future__ import annotations

import ast
import pathlib
import sqlite3
from collections.abc import Iterator
from typing import Any, Final

import pytest
from openpyxl import load_workbook

import core
import db
import export_excel

RAIZ_PROYECTO: Final[pathlib.Path] = pathlib.Path(__file__).resolve().parent.parent
EXPORT_PATH: Final[pathlib.Path] = RAIZ_PROYECTO / "export_excel.py"

TIPO_NORMAL: Final[str] = "Normal (con descuento)"
_METODOS_SQL: Final[frozenset[str]] = frozenset({"execute", "executemany", "executescript"})

#: Encabezado corregido de la hoja "Ventas" aprobado en la desviacion D10.
ENCABEZADO_VENTAS: Final[list[str]] = [
    "Fecha", "Cliente", "Codigo", "Descripcion", "Cantidad", "Precio costo",
    "Precio publico", "Total", "Ganancia", "Pagado", "Saldo",
]

ENCABEZADO_EXISTENCIAS: Final[list[str]] = [
    "Codigo articulo", "Descripcion", "Piezas recibidas", "Piezas vendidas",
    "Piezas disponibles", "Precio unitario costo", "Total pagado real",
    "Valor catalogo total",
]


@pytest.fixture()
def conn() -> Iterator[sqlite3.Connection]:
    """Conexion en memoria con el esquema canonico aplicado."""
    conexion = db.init_db(":memory:")
    try:
        yield conexion
    finally:
        conexion.close()


def _seed_stock(
    conn: sqlite3.Connection,
    *,
    codigo: str,
    descripcion: str = "Articulo",
    piezas: int,
    costo_total: float,
    folio: str = "C001264",
) -> None:
    """Deja `piezas` disponibles del `codigo` en casa, con su costo real."""
    conn.execute(
        "INSERT INTO productos (codigo_articulo, descripcion) VALUES (?, ?)",
        (codigo, descripcion),
    )
    cur = conn.execute("INSERT INTO pedidos (folio_pedido) VALUES (?)", (folio,))
    pedido_id = int(cur.lastrowid)
    conn.execute(
        """
        INSERT INTO pedido_detalle (
            pedido_id, codigo_articulo, ocurrencia, cantidad_solicitada,
            cantidad_surtida, cantidad_asociado, cantidad_casa, cantidad_local,
            precio_que_pagas, valor_total_con_iva, tipo
        ) VALUES (?, ?, 1, ?, ?, 0, ?, 0, ?, ?, ?)
        """,
        (pedido_id, codigo, piezas, piezas, piezas, costo_total, costo_total * 1.5, TIPO_NORMAL),
    )


def _seed_venta_con_pago(conn: sqlite3.Connection) -> None:
    """Siembra dos productos, una venta de dos lineas y un abono parcial."""
    _seed_stock(conn, codigo="11111", descripcion="Sarten", piezas=10, costo_total=1000.0)
    _seed_stock(
        conn, codigo="22222", descripcion="Olla", piezas=8, costo_total=800.0, folio="C001265"
    )
    cliente_id = core.crear_cliente(conn, "Ana Lopez")
    resumen = core.registrar_venta(
        conn,
        cliente_id,
        [
            {"codigo": "11111", "cantidad": 2, "precio_publico": 250.0},
            {"codigo": "22222", "cantidad": 1, "precio_publico": 180.0},
        ],
    )
    core.agregar_pago(conn, "venta_pagos", int(resumen["venta_id"]), "Efectivo", 300.0)


def _exportar(conn: sqlite3.Connection, tmp_path: pathlib.Path) -> pathlib.Path:
    """Exporta a `tmp_path/reporte.xlsx` y devuelve la ruta escrita."""
    destino = tmp_path / "reporte.xlsx"
    export_excel.exportar_a_excel(conn, str(destino))
    return destino


def _valores(hoja: Any) -> list[tuple]:
    """Todas las filas de la hoja como tuplas de valores."""
    return list(hoja.iter_rows(values_only=True))


# --- R1: forma del libro


def test_exportar_a_excel_crea_las_dos_hojas_con_base_sembrada(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    destino = _exportar(conn, tmp_path)

    # Assert
    libro = load_workbook(destino)
    assert libro.sheetnames == ["Existencias", "Ventas"]


def test_exportar_a_excel_devuelve_la_ruta_destino_recibida(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    destino = tmp_path / "salida.xlsx"

    # Act
    devuelto = export_excel.exportar_a_excel(conn, str(destino))

    # Assert
    assert devuelto == str(destino)
    assert destino.exists()


# --- R2: hoja Existencias


def test_hoja_existencias_usa_el_encabezado_del_contrato(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    libro = load_workbook(_exportar(conn, tmp_path))

    # Assert
    assert list(_valores(libro["Existencias"])[0]) == ENCABEZADO_EXISTENCIAS


def test_columnas_existencias_coincide_con_las_claves_de_obtener_existencias(
    conn: sqlite3.Connection,
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    claves_reales = set(core.obtener_existencias(conn)[0])

    # Assert
    assert {clave for _, clave in export_excel.COLUMNAS_EXISTENCIAS} <= claves_reales


def test_hoja_existencias_escribe_una_fila_por_dict_en_el_orden_devuelto(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)
    esperado = [
        tuple(registro[clave] for _, clave in export_excel.COLUMNAS_EXISTENCIAS)
        for registro in core.obtener_existencias(conn)
    ]

    # Act
    filas = _valores(load_workbook(_exportar(conn, tmp_path))["Existencias"])[1:]

    # Assert
    assert len(esperado) == 2
    assert filas == esperado


# --- R3 / D10: hoja Ventas y su contrato


def test_hoja_ventas_usa_el_encabezado_corregido_de_la_desviacion(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    libro = load_workbook(_exportar(conn, tmp_path))

    # Assert
    assert list(_valores(libro["Ventas"])[0]) == ENCABEZADO_VENTAS


def test_columnas_ventas_solo_declara_claves_del_contrato_historial() -> None:
    """Si `CAMPOS_HISTORIAL` cambia, la exportacion falla en vez de vaciarse."""
    # Arrange
    claves = [clave for _, clave in export_excel.COLUMNAS_VENTAS]

    # Act
    ausentes = [clave for clave in claves if clave not in core.CAMPOS_HISTORIAL]

    # Assert
    assert ausentes == []


def test_columnas_ventas_conserva_el_orden_relativo_del_contrato_historial() -> None:
    # Arrange
    claves = [clave for _, clave in export_excel.COLUMNAS_VENTAS]

    # Act
    posiciones = [core.CAMPOS_HISTORIAL.index(clave) for clave in claves]

    # Assert
    assert posiciones == sorted(posiciones)


def test_columnas_ventas_omite_las_columnas_sin_respaldo_en_el_contrato() -> None:
    """`Forma de pago` (N pagos por venta) y `Observaciones` quedan fuera (D10)."""
    # Arrange
    titulos = [titulo for titulo, _ in export_excel.COLUMNAS_VENTAS]

    # Act / Assert
    assert "Forma de pago" not in titulos
    assert "Observaciones" not in titulos


def test_hoja_ventas_puebla_todas_las_celdas_con_una_venta_sembrada(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    """El fallo que evita D10: encabezado correcto con todas las celdas vacias."""
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    filas = _valores(load_workbook(_exportar(conn, tmp_path))["Ventas"])[1:]

    # Assert
    assert len(filas) == 2
    assert all(celda is not None for fila in filas for celda in fila)


def test_hoja_ventas_escribe_una_fila_por_linea_del_historial(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)
    esperado = [
        tuple(registro[clave] for _, clave in export_excel.COLUMNAS_VENTAS)
        for registro in core.obtener_ventas_historial(conn)
    ]

    # Act
    filas = _valores(load_workbook(_exportar(conn, tmp_path))["Ventas"])[1:]

    # Assert
    assert filas == esperado


def test_hoja_ventas_deja_vacia_la_celda_de_una_clave_ausente(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    """Una clave que falte degrada a celda vacia, no lanza (R3)."""
    # Arrange
    destino = tmp_path / "parcial.xlsx"
    libro = export_excel._construir_libro([], [{"fecha": "2026-07-28"}])

    # Act
    libro.save(destino)
    fila = _valores(load_workbook(destino)["Ventas"])[1]

    # Assert
    assert fila[0] == "2026-07-28"
    assert set(fila[1:]) == {None}


# --- R4-R6: estilo


@pytest.mark.parametrize("nombre_hoja", ["Existencias", "Ventas"])
def test_encabezado_lleva_relleno_solido_y_fuente_blanca_en_negrita(
    conn: sqlite3.Connection, tmp_path: pathlib.Path, nombre_hoja: str
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    hoja = load_workbook(_exportar(conn, tmp_path))[nombre_hoja]

    # Assert
    for celda in hoja[1]:
        assert celda.fill.fill_type == "solid"
        assert celda.fill.start_color.rgb[-6:] == "12C1B4"
        assert celda.font.bold is True
        assert celda.font.color.rgb[-6:] == "FFFFFF"


@pytest.mark.parametrize(
    ("nombre_hoja", "columnas_moneda"),
    [
        ("Existencias", export_excel.MONEDA_EXISTENCIAS),
        ("Ventas", export_excel.MONEDA_VENTAS),
    ],
)
def test_columnas_de_dinero_llevan_formato_de_moneda(
    conn: sqlite3.Connection,
    tmp_path: pathlib.Path,
    nombre_hoja: str,
    columnas_moneda: frozenset[str],
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    hoja = load_workbook(_exportar(conn, tmp_path))[nombre_hoja]

    # Assert
    for celdas in hoja.columns:
        esperado = '"$"#,##0.00' if celdas[0].value in columnas_moneda else "General"
        assert [celda.number_format for celda in celdas[1:]] == [esperado] * (len(celdas) - 1)


@pytest.mark.parametrize("nombre_hoja", ["Existencias", "Ventas"])
def test_ancho_de_columna_queda_acotado_entre_el_minimo_y_el_maximo(
    conn: sqlite3.Connection, tmp_path: pathlib.Path, nombre_hoja: str
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)

    # Act
    hoja = load_workbook(_exportar(conn, tmp_path))[nombre_hoja]

    # Assert
    anchos = [dim.width for dim in hoja.column_dimensions.values()]
    assert len(anchos) == hoja.max_column
    assert all(export_excel.ANCHO_MINIMO <= ancho <= export_excel.ANCHO_MAXIMO for ancho in anchos)


def test_ancho_de_columna_se_recorta_al_maximo_con_un_texto_largo(
    tmp_path: pathlib.Path,
) -> None:
    # Arrange
    destino = tmp_path / "ancho.xlsx"
    largo = {"Descripcion": "X" * 200, "fecha": "2026-07-28"}

    # Act
    export_excel._construir_libro([largo], [largo]).save(destino)
    hoja = load_workbook(destino)["Existencias"]

    # Assert
    assert hoja.column_dimensions["B"].width == export_excel.ANCHO_MAXIMO


# --- R7: base vacia


def test_exportar_a_excel_con_base_vacia_genera_solo_los_encabezados(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    destino = tmp_path / "vacio.xlsx"

    # Act
    export_excel.exportar_a_excel(conn, str(destino))
    libro = load_workbook(destino)

    # Assert
    assert libro.sheetnames == ["Existencias", "Ventas"]
    assert _valores(libro["Existencias"]) == [tuple(ENCABEZADO_EXISTENCIAS)]
    assert _valores(libro["Ventas"]) == [tuple(ENCABEZADO_VENTAS)]


# --- R8: el modulo no toca la base


def test_exportar_a_excel_no_escribe_en_la_base(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    _seed_venta_con_pago(conn)
    cambios_previos = conn.total_changes
    historial_previo = core.obtener_ventas_historial(conn)

    # Act
    _exportar(conn, tmp_path)

    # Assert
    assert conn.total_changes == cambios_previos
    assert core.obtener_ventas_historial(conn) == historial_previo


def test_export_excel_no_ejecuta_sql_propio() -> None:
    """El modulo lee via `core`: ningun `execute` debe aparecer en su AST."""
    # Arrange
    arbol = ast.parse(EXPORT_PATH.read_text(encoding="utf-8"))

    # Act
    llamadas = [
        nodo
        for nodo in ast.walk(arbol)
        if isinstance(nodo, ast.Call)
        and isinstance(nodo.func, ast.Attribute)
        and nodo.func.attr in _METODOS_SQL
    ]

    # Assert
    assert llamadas == []


def test_export_excel_lee_solo_por_las_funciones_de_dominio() -> None:
    # Arrange
    fuente = EXPORT_PATH.read_text(encoding="utf-8")

    # Act
    arbol = ast.parse(fuente)
    modulos_dominio = {"core", "core_existencias", "core_historial"}
    atributos = {
        f"{nodo.value.id}.{nodo.attr}"
        for nodo in ast.walk(arbol)
        if isinstance(nodo, ast.Attribute)
        and isinstance(nodo.value, ast.Name)
        and nodo.value.id in modulos_dominio
    }

    # Assert: solo las dos lecturas, y **nunca** la fachada `core` -- importarla
    # desde aqui cerraria el ciclo `core` -> `export_excel` -> `core`.
    assert atributos == {
        "core_existencias.obtener_existencias",
        "core_historial.obtener_ventas_historial",
    }


# --- errores de escritura


def test_exportar_a_excel_envuelve_el_fallo_de_escritura_en_export_error(
    conn: sqlite3.Connection, tmp_path: pathlib.Path
) -> None:
    # Arrange
    destino = tmp_path / "carpeta_inexistente" / "reporte.xlsx"

    # Act / Assert
    with pytest.raises(export_excel.ExportError):
        export_excel.exportar_a_excel(conn, str(destino))


def test_export_error_hereda_de_core_error() -> None:
    # Assert
    assert issubclass(export_excel.ExportError, core.CoreError)
