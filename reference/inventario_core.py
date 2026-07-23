"""
Logica de extraccion y acumulacion de inventario a partir de
remisiones PDF de Betterware, y de registro de ventas.

El registro de ventas ahora se hace desde el programa (gui_inventario.py),
no escribiendo directamente en el Excel. El Excel sigue siendo el
archivo donde queda guardado todo (Movimientos, Existencias, Ventas),
pero ya no tiene formulas ni listas desplegables: los numeros que ves
ahi ya vienen calculados y validados por Python.
"""

import os
import re
from datetime import datetime

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

SHEET_MOV = "Movimientos"
SHEET_STOCK = "Existencias"
SHEET_VENTAS = "Ventas"
SHEET_ASOCIADOS = "Entregas Asociado"
SHEET_DIRECTORIO = "Directorio Asociados"

DIRECTORIO_COLUMNS = ["Nombre", "Telefono", "Notas"]

STOCK_COLUMNS = [
    "Codigo articulo",
    "Descripcion",
    "Piezas recibidas",
    "Piezas vendidas",
    "Piezas disponibles",
    "Precio unitario costo",
    "Total pagado real",
    "Valor catalogo total",
]

VENTAS_COLUMNS = [
    "Fecha",
    "Codigo",
    "Descripcion",
    "Cantidad vendida",
    "Precio asociado",
    "Precio publico",
    "Total",
    "Ganancia",
    "Forma de pago",
    "Observaciones",
]

ASOCIADOS_COLUMNS = [
    "Fecha entrega",
    "Folio de pedido",
    "Codigo",
    "Descripcion",
    "Ocurrencia",
    "Cantidad entregada",
    "Monto que debe pagar",
    "Status",
    "Forma de pago 1",
    "Monto 1",
    "Forma de pago 2",
    "Monto 2",
    "Observaciones",
]

STATUS_ASOCIADO_OPCIONES = ["Pendiente de recoger", "Recogido - no pagado", "Pagado"]
FORMA_PAGO_OPCIONES = ["Efectivo", "Transferencia", "Tarjeta", "Otro"]

STOCK_BAJO_UMBRAL = 3

MOV_COLUMNS = [
    "Fecha registro",
    "Semana",
    "Folio de pedido",
    "Codigo nota",
    "Distribuidora",
    "Nombre asociado",
    "Codigo articulo",
    "Descripcion",
    "Cantidad solicitada",
    "Cantidad surtida",
    "Cantidad Asociado",
    "Cantidad Casa",
    "Cantidad Local",
    "Precio catalogo",
    "Precio con IVA",
    "Precio que pagas",
    "Valor total con IVA",
    "Tipo",
    "Ocurrencia",
    "Archivo origen",
]


class VentaError(Exception):
    """Error de negocio al registrar una venta (ej. stock insuficiente)."""
    pass


def _clean_money(value):
    if value is None:
        return 0.0
    value = str(value).replace("$", "").replace(",", "").strip()
    try:
        return float(value)
    except ValueError:
        return 0.0


def _clean_int(value):
    if value is None:
        return 0
    value = str(value).strip()
    try:
        return int(value)
    except ValueError:
        return 0


def extraer_metadata(pagina_texto):
    meta = {
        "Semana": "",
        "Folio de pedido": "",
        "Codigo nota": "",
        "Distribuidora": "",
        "Nombre asociado": "",
    }

    m = re.search(r"Semana\s+([\d\s\-]+\d)", pagina_texto)
    if m:
        meta["Semana"] = m.group(1).strip()

    m = re.search(r"Folio de pedido\s+(\S+)", pagina_texto)
    if m:
        meta["Folio de pedido"] = m.group(1).strip()

    m = re.search(r"C[oó]digo\s+(\d+)\s+Nombre Asociado", pagina_texto)
    if m:
        meta["Codigo nota"] = m.group(1).strip()

    m = re.search(r"Distribuidora\s+(C\d+[^\n]*)", pagina_texto)
    if m:
        meta["Distribuidora"] = m.group(1).strip()

    m = re.search(r"Nombre Asociado\s+(.+?)\s+Tel[eé]fono Distribuidora", pagina_texto)
    if m:
        meta["Nombre asociado"] = m.group(1).strip()

    return meta


def extraer_productos_de_tabla(tabla):
    productos = []
    for row in tabla:
        # OJO: solo se descartan las celdas que son None (huecos de la
        # tabla). Las celdas vacias pero reales (como "Pag" en blanco
        # cuando un producto no trae numero de pagina) SE CONSERVAN
        # como texto vacio, para no perder columnas y que la fila no
        # se descarte por error.
        limpio = [str(c).strip().replace("\n", " ") if c is not None else None for c in row]
        limpio = [c for c in limpio if c is not None]
        if not limpio:
            continue
        primero = limpio[0]
        if primero in ("Artículo", "Articulo") or primero.startswith("Total"):
            continue
        if not re.match(r"^\d{4,7}$", primero):
            continue

        if len(limpio) == 8:
            codigo, desc, pag, solicitada, surtida, sin_iva, con_iva, valor_total = limpio
            cantidad_surtida = _clean_int(surtida)
            precio_con_iva_unit = _clean_money(con_iva)
            precio_que_pagas = round(precio_con_iva_unit * cantidad_surtida * (1 - 0.18), 2)
            productos.append({
                "Codigo articulo": codigo,
                "Descripcion": desc,
                "Cantidad solicitada": _clean_int(solicitada),
                "Cantidad surtida": cantidad_surtida,
                "Cantidad Asociado": 0,
                "Cantidad Casa": cantidad_surtida,
                "Cantidad Local": 0,
                "Precio catalogo": _clean_money(sin_iva),
                "Precio con IVA": precio_con_iva_unit,
                "Precio que pagas": precio_que_pagas,
                "Valor total con IVA": _clean_money(valor_total),
                "Tipo": "Normal (con descuento)",
            })
        elif len(limpio) == 9:
            codigo, desc, solicitada, surtida, catalogo, paga, gana, con_iva, valor_total = limpio
            cantidad_surtida = _clean_int(surtida)
            productos.append({
                "Codigo articulo": codigo,
                "Descripcion": desc,
                "Cantidad solicitada": _clean_int(solicitada),
                "Cantidad surtida": cantidad_surtida,
                "Cantidad Asociado": 0,
                "Cantidad Casa": cantidad_surtida,
                "Cantidad Local": 0,
                "Precio catalogo": _clean_money(catalogo),
                "Precio con IVA": _clean_money(con_iva),
                "Precio que pagas": _clean_money(paga),
                "Valor total con IVA": _clean_money(valor_total),
                "Tipo": "Sin descuento",
            })
    return productos


def procesar_pdf(ruta_pdf):
    """Procesa un PDF que puede traer un pedido o varios (uno por
    pagina). Cada pagina se revisa por separado: si trae su propio
    encabezado (Semana/Folio de pedido), se toma como un pedido nuevo;
    si no trae encabezado, se asume que es continuacion del pedido de
    la pagina anterior (un mismo pedido repartido en varias paginas)."""
    filas = []
    meta_actual = {
        "Semana": "",
        "Folio de pedido": "",
        "Codigo nota": "",
        "Distribuidora": "",
        "Nombre asociado": "",
    }
    with pdfplumber.open(ruta_pdf) as pdf:
        for num_pagina, pagina in enumerate(pdf.pages, start=1):
            texto_pagina = pagina.extract_text() or ""
            meta_pagina = extraer_metadata(texto_pagina)

            if meta_pagina["Folio de pedido"]:
                meta_actual = meta_pagina

            for tabla in pagina.extract_tables():
                if not tabla:
                    continue
                primera_fila = [str(c).strip() if c else "" for c in tabla[0]]
                if not any("Art" in c for c in primera_fila):
                    continue
                for prod in extraer_productos_de_tabla(tabla):
                    fila = {
                        "Fecha registro": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "Semana": meta_actual["Semana"],
                        "Folio de pedido": meta_actual["Folio de pedido"],
                        "Codigo nota": meta_actual["Codigo nota"],
                        "Distribuidora": meta_actual["Distribuidora"],
                        "Nombre asociado": meta_actual["Nombre asociado"],
                        "Archivo origen": f"{os.path.basename(ruta_pdf)} (pag. {num_pagina})",
                    }
                    fila.update(prod)
                    filas.append(fila)

    # Si un mismo producto aparece mas de una vez en el mismo pedido
    # (por ejemplo, se agrego dos veces al carrito), cada aparicion se
    # numera. Esto evita que, al guardar, se confunda con un duplicado
    # y se pierda una de las piezas.
    contador_ocurrencias = {}
    for fila in filas:
        clave = (fila["Folio de pedido"], fila["Codigo articulo"], fila["Tipo"])
        contador_ocurrencias[clave] = contador_ocurrencias.get(clave, 0) + 1
        fila["Ocurrencia"] = contador_ocurrencias[clave]

    return filas


def cargar_existente(ruta_excel):
    if os.path.exists(ruta_excel):
        try:
            df = pd.read_excel(ruta_excel, sheet_name=SHEET_MOV)
            df["Codigo articulo"] = df["Codigo articulo"].astype(str).str.strip()
            # Compatibilidad con archivos generados antes de tener la
            # repartición Asociado/Casa/Local: se asume que todo se
            # quedo en "Casa" (como era el comportamiento anterior).
            if "Cantidad Casa" not in df.columns:
                df["Cantidad Asociado"] = 0
                df["Cantidad Casa"] = df["Cantidad surtida"]
                df["Cantidad Local"] = 0
            for col in ("Cantidad Asociado", "Cantidad Casa", "Cantidad Local"):
                df[col] = df[col].fillna(0).astype(int)
            return df
        except Exception:
            pass
    return pd.DataFrame(columns=MOV_COLUMNS)


def leer_ventas(ruta_excel):
    """Lee la hoja Ventas tal cual (ya no tiene formulas: son valores
    normales, asi que se puede leer directo con pandas)."""
    if os.path.exists(ruta_excel):
        try:
            hojas = pd.ExcelFile(ruta_excel).sheet_names
            if SHEET_VENTAS in hojas:
                df = pd.read_excel(ruta_excel, sheet_name=SHEET_VENTAS)
                if "Codigo" in df.columns:
                    df["Codigo"] = df["Codigo"].astype(str).str.strip()
                    return df[[c for c in VENTAS_COLUMNS if c in df.columns]]
                # Formato antiguo (columna "Producto" combinada o con formulas):
                # se migra usando los ultimos valores que haya calculado Excel.
                return _migrar_ventas_formato_antiguo(ruta_excel)
        except Exception:
            pass
    return pd.DataFrame(columns=VENTAS_COLUMNS)


def _migrar_ventas_formato_antiguo(ruta_excel):
    """Convierte una hoja Ventas de una version anterior (con columna
    'Producto' tipo 'codigo - nombre', y formulas para Codigo/Precio
    asociado/Total/Ganancia) al formato nuevo, usando los ultimos
    valores que haya calculado Excel."""
    try:
        wb = load_workbook(ruta_excel, data_only=True)
    except Exception:
        return pd.DataFrame(columns=VENTAS_COLUMNS)
    if SHEET_VENTAS not in wb.sheetnames:
        return pd.DataFrame(columns=VENTAS_COLUMNS)

    ws = wb[SHEET_VENTAS]
    encabezados = [c.value for c in ws[1]]
    idx = {nombre: i for i, nombre in enumerate(encabezados) if nombre}

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        producto = row[idx["Producto"]] if "Producto" in idx else None
        if not producto:
            continue
        codigo = row[idx["Codigo"]] if "Codigo" in idx else None
        if not codigo:
            m = re.match(r"^\s*(\d{4,7})", str(producto))
            codigo = m.group(1) if m else ""
        descripcion = str(producto).split(" - ", 1)[-1].strip() if " - " in str(producto) else str(producto)

        filas.append({
            "Fecha": row[idx["Fecha"]] if "Fecha" in idx else None,
            "Codigo": str(codigo).strip(),
            "Descripcion": descripcion,
            "Cantidad vendida": row[idx["Cantidad vendida"]] if "Cantidad vendida" in idx else 0,
            "Precio asociado": row[idx["Precio asociado"]] if "Precio asociado" in idx else 0,
            "Precio publico": row[idx["Precio publico"]] if "Precio publico" in idx else 0,
            "Total": row[idx["Total"]] if "Total" in idx else 0,
            "Ganancia": row[idx["Ganancia"]] if "Ganancia" in idx else 0,
            "Forma de pago": row[idx["Forma de pago"]] if "Forma de pago" in idx else "",
            "Observaciones": row[idx["Observaciones"]] if "Observaciones" in idx else "",
        })
    wb.close()
    return pd.DataFrame(filas, columns=VENTAS_COLUMNS)


def construir_existencias(df_mov, df_ventas=None):
    if df_mov.empty:
        return pd.DataFrame(columns=STOCK_COLUMNS)

    df = df_mov.copy()
    # Solo cuenta como "tu" stock lo que se quedo en Casa o Local; lo
    # que se lleva el Asociado no se contabiliza como tuyo.
    df["Cantidad stock propio"] = df["Cantidad Casa"].fillna(0) + df["Cantidad Local"].fillna(0)

    def _proporcional(row, campo):
        total_surtido = row["Cantidad surtida"]
        if not total_surtido:
            return 0.0
        return row[campo] * (row["Cantidad stock propio"] / total_surtido)

    df["Costo propio"] = df.apply(lambda r: _proporcional(r, "Precio que pagas"), axis=1)
    df["Valor catalogo propio"] = df.apply(lambda r: _proporcional(r, "Valor total con IVA"), axis=1)

    resumen = (
        df.groupby(["Codigo articulo", "Descripcion"], as_index=False)
        .agg(
            **{
                "Piezas recibidas": ("Cantidad stock propio", "sum"),
                "Total pagado real": ("Costo propio", "sum"),
                "Valor catalogo total": ("Valor catalogo propio", "sum"),
            }
        )
        .sort_values("Codigo articulo")
        .reset_index(drop=True)
    )
    resumen["Piezas recibidas"] = resumen["Piezas recibidas"].round().astype(int)
    resumen["Total pagado real"] = resumen["Total pagado real"].round(2)
    resumen["Valor catalogo total"] = resumen["Valor catalogo total"].round(2)

    if df_ventas is not None and not df_ventas.empty:
        vendidas = (
            df_ventas.groupby("Codigo", as_index=False)["Cantidad vendida"].sum()
            .rename(columns={"Codigo": "Codigo articulo", "Cantidad vendida": "Piezas vendidas"})
        )
        resumen = resumen.merge(vendidas, on="Codigo articulo", how="left")
        resumen["Piezas vendidas"] = resumen["Piezas vendidas"].fillna(0).astype(int)
    else:
        resumen["Piezas vendidas"] = 0

    resumen["Piezas disponibles"] = resumen["Piezas recibidas"] - resumen["Piezas vendidas"]
    resumen["Precio unitario costo"] = resumen.apply(
        lambda r: round(r["Total pagado real"] / r["Piezas recibidas"], 2) if r["Piezas recibidas"] else 0.0,
        axis=1,
    )
    resumen = resumen[STOCK_COLUMNS]
    return resumen


def leer_entregas_asociado(ruta_excel):
    """Lee la hoja de entregas a Asociado tal cual esta guardada (para
    conservar el Status y las formas de pago que ya se hayan capturado)."""
    if os.path.exists(ruta_excel):
        try:
            hojas = pd.ExcelFile(ruta_excel).sheet_names
            if SHEET_ASOCIADOS in hojas:
                df = pd.read_excel(ruta_excel, sheet_name=SHEET_ASOCIADOS)
                for col in ("Codigo", "Folio de pedido"):
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.strip()
                if "Ocurrencia" in df.columns:
                    df["Ocurrencia"] = df["Ocurrencia"].fillna(1).astype(int)
                return df[[c for c in ASOCIADOS_COLUMNS if c in df.columns]]
        except Exception:
            pass
    return pd.DataFrame(columns=ASOCIADOS_COLUMNS)


def construir_entregas_asociado(df_mov, entregas_previas):
    """A partir de Movimientos, crea una fila en 'Entregas Asociado'
    por cada renglon que tenga piezas marcadas como 'Asociado', y
    conserva el Status/pagos que ya se hubieran capturado antes."""
    if entregas_previas is None:
        entregas_previas = pd.DataFrame(columns=ASOCIADOS_COLUMNS)

    claves_existentes = set(
        zip(
            entregas_previas.get("Folio de pedido", pd.Series(dtype=str)),
            entregas_previas.get("Codigo", pd.Series(dtype=str)),
            entregas_previas.get("Ocurrencia", pd.Series(dtype=int)),
        )
    )

    filas_nuevas = []
    if not df_mov.empty:
        for _, fila in df_mov.iterrows():
            cantidad_asociado = int(fila.get("Cantidad Asociado", 0) or 0)
            if cantidad_asociado <= 0:
                continue
            clave = (str(fila["Folio de pedido"]), str(fila["Codigo articulo"]), int(fila.get("Ocurrencia", 1)))
            if clave in claves_existentes:
                continue  # ya existe (se preserva tal cual con su Status/pagos)

            cantidad_surtida = fila.get("Cantidad surtida", 0) or 0
            precio_que_pagas = fila.get("Precio que pagas", 0) or 0
            monto_debe = round((precio_que_pagas / cantidad_surtida) * cantidad_asociado, 2) if cantidad_surtida else 0.0

            filas_nuevas.append({
                "Fecha entrega": fila.get("Fecha registro"),
                "Folio de pedido": clave[0],
                "Codigo": clave[1],
                "Descripcion": fila.get("Descripcion"),
                "Ocurrencia": clave[2],
                "Cantidad entregada": cantidad_asociado,
                "Monto que debe pagar": monto_debe,
                "Status": STATUS_ASOCIADO_OPCIONES[0],
                "Forma de pago 1": "",
                "Monto 1": None,
                "Forma de pago 2": "",
                "Monto 2": None,
                "Observaciones": "",
            })

    if filas_nuevas:
        df_nuevas = pd.DataFrame(filas_nuevas, columns=ASOCIADOS_COLUMNS)
        return pd.concat([entregas_previas, df_nuevas], ignore_index=True)
    return entregas_previas


def aplicar_listas_asociado(ws, ultima_fila):
    """Listas desplegables simples (sin formulas) para Status y Forma
    de pago en la hoja de entregas a Asociado. Se dejan como respaldo,
    pero lo normal es administrar todo esto desde el programa."""
    if ultima_fila < 2:
        return
    dv_status = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_ASOCIADO_OPCIONES) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"H2:H{ultima_fila}")  # columna "Status"

    for columna_letra in ("I", "K"):  # Forma de pago 1 y Forma de pago 2
        dv_pago = DataValidation(
            type="list",
            formula1='"' + ",".join(FORMA_PAGO_OPCIONES) + '"',
            allow_blank=True,
        )
        ws.add_data_validation(dv_pago)
        dv_pago.add(f"{columna_letra}2:{columna_letra}{ultima_fila}")


# ---------------------------------------------------------------------
# Directorio de Asociados (nombre, telefono, notas) + link de WhatsApp
# ---------------------------------------------------------------------

def leer_directorio_asociados(ruta_excel):
    if os.path.exists(ruta_excel):
        try:
            hojas = pd.ExcelFile(ruta_excel).sheet_names
            if SHEET_DIRECTORIO in hojas:
                df = pd.read_excel(ruta_excel, sheet_name=SHEET_DIRECTORIO)
                return df[[c for c in DIRECTORIO_COLUMNS if c in df.columns]]
        except Exception:
            pass
    return pd.DataFrame(columns=DIRECTORIO_COLUMNS)


def limpiar_telefono(telefono):
    """Deja solo digitos y antepone codigo de pais (52, Mexico) si
    parece un numero local de 10 digitos."""
    digitos = re.sub(r"\D", "", str(telefono or ""))
    if len(digitos) == 10:
        digitos = "52" + digitos
    return digitos


def link_whatsapp(telefono, mensaje=""):
    """Regresa el link de wa.me listo para abrir en el navegador con
    un mensaje precargado (opcional)."""
    numero = limpiar_telefono(telefono)
    if not numero:
        return None
    if mensaje:
        from urllib.parse import quote
        return f"https://wa.me/{numero}?text={quote(mensaje)}"
    return f"https://wa.me/{numero}"


def agregar_asociado(ruta_excel, nombre, telefono, notas=""):
    if not nombre or not nombre.strip():
        raise ValueError("El nombre del asociado no puede estar vacio.")
    df = leer_directorio_asociados(ruta_excel)
    nueva_fila = {"Nombre": nombre.strip(), "Telefono": str(telefono or "").strip(), "Notas": notas or ""}
    df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)
    _guardar_hoja_directorio(ruta_excel, df)
    return df


def editar_asociado(ruta_excel, indice, nombre=None, telefono=None, notas=None):
    df = leer_directorio_asociados(ruta_excel)
    if indice < 0 or indice >= len(df):
        raise ValueError("No se encontro ese asociado.")
    if nombre is not None:
        df.at[indice, "Nombre"] = nombre.strip()
    if telefono is not None:
        df.at[indice, "Telefono"] = str(telefono).strip()
    if notas is not None:
        df.at[indice, "Notas"] = notas
    _guardar_hoja_directorio(ruta_excel, df)
    return df


def eliminar_asociado(ruta_excel, indice):
    df = leer_directorio_asociados(ruta_excel)
    if indice < 0 or indice >= len(df):
        raise ValueError("No se encontro ese asociado.")
    df = df.drop(index=indice).reset_index(drop=True)
    _guardar_hoja_directorio(ruta_excel, df)
    return df


def _guardar_hoja_directorio(ruta_excel, df_directorio):
    """Guarda solo la hoja de directorio, preservando todo lo demas."""
    df_mov = cargar_existente(ruta_excel)
    df_ventas = leer_ventas(ruta_excel)
    df_asociados = leer_entregas_asociado(ruta_excel)
    df_existencias = construir_existencias(df_mov, df_ventas)
    _guardar_excel_completo(ruta_excel, df_mov, df_existencias, df_ventas, df_asociados, df_directorio)


# ---------------------------------------------------------------------
# Gestion de entregas a Asociado desde el programa (status y pagos)
# ---------------------------------------------------------------------

def obtener_entregas_asociado(ruta_excel):
    """Regresa la lista de entregas a Asociado como dicts, con un
    indice (posicion en la tabla) para poder actualizarlas despues."""
    df = leer_entregas_asociado(ruta_excel)
    registros = df.to_dict("records")
    for i, r in enumerate(registros):
        r["_indice"] = i
    return registros


def actualizar_entrega_asociado(ruta_excel, indice, status=None, forma_pago_1=None, monto_1=None,
                                 forma_pago_2=None, monto_2=None, observaciones=None):
    """Actualiza el Status y/o los pagos de una entrega a Asociado ya
    existente (identificada por su indice en la tabla)."""
    df_entregas = leer_entregas_asociado(ruta_excel)
    if indice < 0 or indice >= len(df_entregas):
        raise ValueError("No se encontro esa entrega.")

    # Si una columna llego vacia (todo NaN), pandas la infiere como
    # numerica; hay que pasarla a texto/objeto antes de poder meterle
    # strings como "Efectivo" o "Pagado".
    for col in ("Status", "Forma de pago 1", "Forma de pago 2", "Observaciones"):
        if col in df_entregas.columns and df_entregas[col].dtype != object:
            df_entregas[col] = df_entregas[col].astype(object)
    for col in ("Monto 1", "Monto 2"):
        if col in df_entregas.columns:
            df_entregas[col] = df_entregas[col].astype(float)

    if status is not None:
        df_entregas.at[indice, "Status"] = status
    if forma_pago_1 is not None:
        df_entregas.at[indice, "Forma de pago 1"] = forma_pago_1
    if monto_1 is not None:
        df_entregas.at[indice, "Monto 1"] = float(monto_1)
    if forma_pago_2 is not None:
        df_entregas.at[indice, "Forma de pago 2"] = forma_pago_2
    if monto_2 is not None:
        df_entregas.at[indice, "Monto 2"] = float(monto_2)
    if observaciones is not None:
        df_entregas.at[indice, "Observaciones"] = observaciones

    df_mov = cargar_existente(ruta_excel)
    df_ventas = leer_ventas(ruta_excel)
    df_existencias = construir_existencias(df_mov, df_ventas)
    df_directorio = leer_directorio_asociados(ruta_excel)
    _guardar_excel_completo(ruta_excel, df_mov, df_existencias, df_ventas, df_entregas, df_directorio)
    return df_entregas.loc[indice].to_dict()


# ---------------------------------------------------------------------
# Datos para el dashboard y las vistas de Pedidos/Ventas (filtrables)
# ---------------------------------------------------------------------

def obtener_movimientos(ruta_excel):
    """Lista de todos los renglones de pedidos (Movimientos), para la
    pantalla de 'Pedidos' con filtros."""
    df = cargar_existente(ruta_excel)
    return df.to_dict("records")


def obtener_ventas_historial(ruta_excel):
    """Lista de todas las ventas registradas, para la pantalla de
    'Ventas'."""
    df = leer_ventas(ruta_excel)
    return df.to_dict("records")


def obtener_resumen_dashboard(ruta_excel):
    """Numeros clave para la pantalla principal (Dashboard)."""
    df_mov = cargar_existente(ruta_excel)
    df_ventas = leer_ventas(ruta_excel)
    df_existencias = construir_existencias(df_mov, df_ventas)
    df_entregas = leer_entregas_asociado(ruta_excel)

    resumen = {
        "productos_distintos": int(len(df_existencias)),
        "piezas_disponibles": int(df_existencias["Piezas disponibles"].sum()) if not df_existencias.empty else 0,
        "valor_inventario_costo": round(
            (df_existencias["Piezas disponibles"] * df_existencias["Precio unitario costo"]).sum(), 2
        ) if not df_existencias.empty else 0.0,
        "productos_bajo_stock": (
            df_existencias[df_existencias["Piezas disponibles"] <= STOCK_BAJO_UMBRAL][
                ["Codigo articulo", "Descripcion", "Piezas disponibles"]
            ].to_dict("records")
            if not df_existencias.empty else []
        ),
        "num_ventas": int(len(df_ventas)),
        "total_vendido": round(df_ventas["Total"].sum(), 2) if not df_ventas.empty else 0.0,
        "ganancia_total": round(df_ventas["Ganancia"].sum(), 2) if not df_ventas.empty else 0.0,
        "num_pedidos_distintos": int(df_mov["Folio de pedido"].nunique()) if not df_mov.empty else 0,
        "entregas_pendientes_cobro": (
            int((df_entregas["Status"] != "Pagado").sum()) if not df_entregas.empty else 0
        ),
        "monto_pendiente_asociados": (
            round(df_entregas.loc[df_entregas["Status"] != "Pagado", "Monto que debe pagar"].sum(), 2)
            if not df_entregas.empty else 0.0
        ),
    }
    return resumen


def formatear_hoja(ws, es_moneda_cols=None):
    es_moneda_cols = es_moneda_cols or []
    header_fill = PatternFill(start_color="12C1B4", end_color="12C1B4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        header_val = col_cells[0].value
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)
        if header_val in es_moneda_cols:
            for c in col_cells[1:]:
                c.number_format = '"$"#,##0.00'


def aplicar_alerta_stock_bajo(ws, ultima_fila):
    if ultima_fila < 2:
        return
    rango = f"E2:E{ultima_fila}"
    rojo_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    rojo_font = Font(color="9C0006")
    regla = CellIsRule(operator="lessThanOrEqual", formula=[str(STOCK_BAJO_UMBRAL)], fill=rojo_fill, font=rojo_font)
    ws.conditional_formatting.add(rango, regla)


def _guardar_excel_completo(ruta_excel, df_mov, df_existencias, df_ventas, df_asociados=None, df_directorio=None):
    if df_asociados is None:
        df_asociados = pd.DataFrame(columns=ASOCIADOS_COLUMNS)
    if df_directorio is None:
        df_directorio = pd.DataFrame(columns=DIRECTORIO_COLUMNS)

    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        df_mov.to_excel(writer, sheet_name=SHEET_MOV, index=False)
        df_existencias.to_excel(writer, sheet_name=SHEET_STOCK, index=False)
        df_ventas.to_excel(writer, sheet_name=SHEET_VENTAS, index=False)
        df_asociados.to_excel(writer, sheet_name=SHEET_ASOCIADOS, index=False)
        df_directorio.to_excel(writer, sheet_name=SHEET_DIRECTORIO, index=False)

    wb = load_workbook(ruta_excel)
    formatear_hoja(wb[SHEET_MOV], es_moneda_cols=["Precio catalogo", "Precio con IVA", "Precio que pagas", "Valor total con IVA"])

    ws_stock = wb[SHEET_STOCK]
    formatear_hoja(ws_stock, es_moneda_cols=["Precio unitario costo", "Total pagado real", "Valor catalogo total"])
    aplicar_alerta_stock_bajo(ws_stock, len(df_existencias) + 1)

    formatear_hoja(wb[SHEET_VENTAS], es_moneda_cols=["Precio asociado", "Precio publico", "Total", "Ganancia"])

    ws_asociados = wb[SHEET_ASOCIADOS]
    formatear_hoja(ws_asociados, es_moneda_cols=["Monto que debe pagar", "Monto 1", "Monto 2"])
    aplicar_listas_asociado(ws_asociados, len(df_asociados) + 1)

    formatear_hoja(wb[SHEET_DIRECTORIO])

    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta_excel)


def actualizar_excel_maestro(nuevas_filas, ruta_excel):
    """Se llama cuando se procesan PDFs nuevos: agrega los productos
    recibidos, conservando las ventas ya registradas."""
    df_ventas = leer_ventas(ruta_excel)
    entregas_previas = leer_entregas_asociado(ruta_excel)

    df_nuevo = pd.DataFrame(nuevas_filas, columns=MOV_COLUMNS)
    df_existente = cargar_existente(ruta_excel)

    df_total = pd.concat([df_existente, df_nuevo], ignore_index=True)
    df_total["Codigo articulo"] = df_total["Codigo articulo"].astype(str).str.strip()
    df_total["Codigo nota"] = df_total["Codigo nota"].astype(str).str.strip()
    if "Ocurrencia" not in df_total.columns:
        df_total["Ocurrencia"] = 1
    df_total["Ocurrencia"] = df_total["Ocurrencia"].fillna(1).astype(int)
    df_total = df_total.drop_duplicates(
        subset=["Folio de pedido", "Codigo articulo", "Tipo", "Ocurrencia"], keep="last"
    )

    df_existencias = construir_existencias(df_total, df_ventas)
    df_asociados = construir_entregas_asociado(df_total, entregas_previas)
    df_directorio = leer_directorio_asociados(ruta_excel)
    _guardar_excel_completo(ruta_excel, df_total, df_existencias, df_ventas, df_asociados, df_directorio)

    return df_total, df_existencias


def procesar_varios_pdfs(rutas_pdf, ruta_excel):
    """Procesa una lista de rutas de PDF y actualiza el Excel maestro
    directamente, SIN previsualizar (se usa desde la linea de
    comandos). La interfaz grafica usa en su lugar
    preparar_filas_desde_pdfs() + confirmar_carga() para poder
    mostrar una vista previa editable antes de guardar."""
    todas_las_filas, errores = preparar_filas_desde_pdfs(rutas_pdf)
    if todas_las_filas:
        actualizar_excel_maestro(todas_las_filas, ruta_excel)
    return len(todas_las_filas), errores


def preparar_filas_desde_pdfs(rutas_pdf):
    """Extrae los productos de uno o varios PDF SIN guardarlos todavia
    en el Excel, para poder mostrarlos en una vista previa editable
    antes de confirmar la carga. Regresa (lista_de_filas, errores)."""
    todas_las_filas = []
    errores = []
    for ruta in rutas_pdf:
        try:
            filas = procesar_pdf(ruta)
            if not filas:
                errores.append(f"{os.path.basename(ruta)}: no se encontraron productos (revisa que sea una remision valida)")
            todas_las_filas.extend(filas)
        except Exception as e:
            errores.append(f"{os.path.basename(ruta)}: error al leer el archivo ({e})")
    return todas_las_filas, errores


def confirmar_carga(filas, ruta_excel):
    """Guarda en el Excel maestro las filas ya revisadas/editadas por
    el usuario en la vista previa."""
    if not filas:
        return None, None
    return actualizar_excel_maestro(filas, ruta_excel)


# ---------------------------------------------------------------------
# Registro de ventas (ahora se maneja desde el programa, no en Excel)
# ---------------------------------------------------------------------

def obtener_catalogo(ruta_excel):
    """Regresa la lista de productos disponibles para vender, para
    mostrar en el buscador de la ventana de Ventas."""
    df_mov = cargar_existente(ruta_excel)
    if df_mov.empty:
        return []
    df_ventas = leer_ventas(ruta_excel)
    df_existencias = construir_existencias(df_mov, df_ventas)
    return df_existencias.to_dict("records")


def registrar_venta(ruta_excel, codigo, cantidad, precio_publico, forma_pago, observaciones="", fecha=None):
    """Registra una venta nueva. Lanza VentaError si el producto no
    existe o si no hay suficiente inventario disponible. Regresa un
    dict con el resumen de la venta ya registrada."""
    codigo = str(codigo).strip()
    if not codigo:
        raise VentaError("Selecciona un producto.")
    if cantidad is None or cantidad <= 0:
        raise VentaError("La cantidad vendida debe ser mayor a cero.")

    df_mov = cargar_existente(ruta_excel)
    if df_mov.empty:
        raise VentaError("Todavia no hay productos en el inventario.")

    df_ventas = leer_ventas(ruta_excel)
    df_existencias = construir_existencias(df_mov, df_ventas)

    fila_producto = df_existencias[df_existencias["Codigo articulo"] == codigo]
    if fila_producto.empty:
        raise VentaError("No se encontro ese producto en el inventario.")

    fila_producto = fila_producto.iloc[0]
    disponibles = int(fila_producto["Piezas disponibles"])
    if cantidad > disponibles:
        raise VentaError(
            f"Solo hay {disponibles} pieza(s) disponibles de '{fila_producto['Descripcion']}'. "
            f"No se puede vender {cantidad}."
        )

    precio_asociado = float(fila_producto["Precio unitario costo"])
    precio_publico = float(precio_publico)
    total = round(cantidad * precio_publico, 2)
    ganancia = round(total - (cantidad * precio_asociado), 2)

    nueva_venta = {
        "Fecha": fecha or datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Codigo": codigo,
        "Descripcion": fila_producto["Descripcion"],
        "Cantidad vendida": int(cantidad),
        "Precio asociado": precio_asociado,
        "Precio publico": precio_publico,
        "Total": total,
        "Ganancia": ganancia,
        "Forma de pago": forma_pago or "",
        "Observaciones": observaciones or "",
    }

    df_ventas = pd.concat([df_ventas, pd.DataFrame([nueva_venta])], ignore_index=True)
    df_existencias_final = construir_existencias(df_mov, df_ventas)
    df_asociados = leer_entregas_asociado(ruta_excel)
    df_directorio = leer_directorio_asociados(ruta_excel)
    _guardar_excel_completo(ruta_excel, df_mov, df_existencias_final, df_ventas, df_asociados, df_directorio)

    disponibles_restantes = disponibles - int(cantidad)
    return {
        "descripcion": fila_producto["Descripcion"],
        "total": total,
        "ganancia": ganancia,
        "disponibles_restantes": disponibles_restantes,
    }
